import os
import openpyxl
import random
from datetime import datetime


def read_excel_column(sheet_name, column_index):
    # 获取项目根目录
    current_directory = os.path.dirname(os.path.abspath(__file__))
    file_path = os.path.join(current_directory, '2026届13班座位表源.xlsx')

    # 加载 Excel 文件
    workbook = openpyxl.load_workbook(file_path)

    # 检查工作表是否存在
    if sheet_name not in workbook.sheetnames:
        raise KeyError(f"Worksheet '{sheet_name}' does not exist in the workbook.")

    # 获取指定的工作表
    sheet = workbook[sheet_name]

    # 获取最大行数
    max_row = sheet.max_row

    # 读取指定列的数据
    column_data = [sheet.cell(row=row, column=column_index).value for row in range(1, max_row + 1)]

    # 生成一个随机行号
    random_row = random.randint(1, max_row)

    # 读取指定行的第二列数据
    cell_value = sheet.cell(row=random_row, column=column_index).value

    return workbook, file_path, cell_value, random_row, column_data


def assign_data_to_target_sheet(workbook, file_path, cell_value, target_sheet_name, target_row, target_column):
    # 获取目标工作表
    if target_sheet_name not in workbook.sheetnames:
        raise KeyError(f"Worksheet '{target_sheet_name}' does not exist in the workbook.")
    target_sheet = workbook[target_sheet_name]

    # 分配数据
    target_sheet.cell(row=target_row, column=target_column).value = cell_value

    # 保存更改后的 Excel 文件
    workbook.save(file_path)


def update_output_sheet(workbook, file_path, original_sheet_name, output_sheet_name, column_index):
    # 获取原始工作表
    original_sheet = workbook[original_sheet_name]

    # 获取最大行数
    max_row = original_sheet.max_row

    # 读取指定列的数据
    original_data = [original_sheet.cell(row=row, column=column_index).value for row in range(1, max_row + 1)]

    # 获取目标工作表
    if output_sheet_name not in workbook.sheetnames:
        workbook.create_sheet(output_sheet_name)
    output_sheet = workbook[output_sheet_name]

    # 清空输出工作表
    for row in output_sheet.iter_rows(min_row=1, max_col=1, max_row=max_row):
        for cell in row:
            cell.value = None

    # 将剩余的数据写入输出工作表
    for index, value in enumerate(original_data, 1):
        output_sheet.cell(row=index, column=1).value = value

    # 保存更改后的 Excel 文件
    workbook.save(file_path)


def assign_output_data_to_target_sheet(workbook, file_path, output_sheet_name, target_sheet_name, target_columns,
                                       target_rows):
    # 获取输出工作表
    if output_sheet_name not in workbook.sheetnames:
        raise KeyError(f"Worksheet '{output_sheet_name}' does not exist in the workbook.")
    output_sheet = workbook[output_sheet_name]

    # 获取目标工作表
    if target_sheet_name not in workbook.sheetnames:
        raise KeyError(f"Worksheet '{target_sheet_name}' does not exist in the workbook.")
    target_sheet = workbook[target_sheet_name]

    # 获取输出工作表的最大行数
    max_row = output_sheet.max_row

    # 读取输出工作表的数据
    output_data = [output_sheet.cell(row=row, column=1).value for row in range(1, max_row + 1)]

    # 分配数据到目标工作表
    for i, value in enumerate(output_data):
        row_index = i % len(target_rows) + 1
        col_index = (i // len(target_rows)) % len(target_columns) + 1
        target_sheet.cell(row=row_index, column=target_columns[col_index - 1]).value = value

    # 保存更改后的 Excel 文件
    workbook.save(file_path)


def assign_partner_data_to_output_sheet(workbook, file_path, partner_sheets, output_sheet_name, total_rows):
    # 获取目标工作表
    if output_sheet_name not in workbook.sheetnames:
        raise KeyError(f"Worksheet '{output_sheet_name}' does not exist in the workbook.")
    output_sheet = workbook[output_sheet_name]

    # 获取所有同桌工作表的数据
    all_partner_data = []
    for sheet_name in partner_sheets:
        if sheet_name not in workbook.sheetnames:
            raise KeyError(f"Worksheet '{sheet_name}' does not exist in the workbook.")
        sheet = workbook[sheet_name]
        # 读取所有列的数据
        all_partner_data.extend(
            [[sheet.cell(row=row, column=col).value for col in range(1, sheet.max_column + 1)] for row in
             range(1, sheet.max_row + 1)])

    # 打乱数据
    random.shuffle(all_partner_data)

    # 分配数据到目标工作表
    for i in range(total_rows):
        row_data = all_partner_data[i]
        for j, value in enumerate(row_data):
            output_sheet.cell(row=i + 1, column=j + 1).value = value

    # 保存更改后的 Excel 文件
    workbook.save(file_path)


def copy_seats_to_output_sheet(workbook, file_path, seats_sheet_name, output_sheet_name):
    # 获取“座位表（按同桌）”工作表
    seats_sheet = workbook[seats_sheet_name]

    # 获取“输出座位表”工作表
    output_sheet = workbook[output_sheet_name]

    # 定义数据映射关系
    mapping = [
        (1, 2, 1, 1),  # 添加这一行
        (1, 3, 1, 2),
        (2, 2, 2, 1), (2, 3, 2, 2),
        (3, 2, 3, 1), (3, 3, 3, 2),
        (4, 2, 4, 1), (4, 3, 4, 2),
        (5, 2, 5, 1), (5, 3, 5, 2),
        (6, 2, 6, 1), (6, 3, 6, 2),
        (7, 2, 7, 1), (7, 3, 7, 2),

        (8, 2, 1, 4), (8, 3, 1, 5),
        (9, 2, 2, 4), (9, 3, 2, 5),
        (10, 2, 3, 4), (10, 3, 3, 5),
        (11, 2, 4, 4), (11, 3, 4, 5),
        (12, 2, 5, 4), (12, 3, 5, 5),
        (13, 2, 6, 4), (13, 3, 6, 5),
        (14, 2, 7, 4), (14, 3, 7, 5),

        (15, 2, 1, 7), (15, 3, 1, 8),
        (16, 2, 2, 7), (16, 3, 2, 8),
        (17, 2, 3, 7), (17, 3, 3, 8),
        (18, 2, 5, 7), (18, 3, 5, 8),
        (19, 2, 6, 7), (19, 3, 6, 8),
        (20, 2, 7, 7), (20, 3, 7, 8),

        (21, 2, 1, 10), (21, 3, 1, 11),
        (22, 2, 2, 10), (22, 3, 2, 11),
        (23, 2, 3, 10), (23, 3, 3, 11),
        (24, 2, 4, 10), (24, 3, 4, 11),
        (25, 2, 5, 10), (25, 3, 5, 11),
        (26, 2, 6, 10), (26, 3, 6, 11),
    ]

    # 复制数据
    for src_row, src_col, dst_row, dst_col in mapping:
        value = seats_sheet.cell(row=src_row, column=src_col).value
        output_sheet.cell(row=dst_row, column=dst_col).value = value

    # 保存更改后的 Excel 文件
    workbook.save(file_path)


def create_new_file_with_output_sheet(workbook, source_sheet_name):
    # 获取当前日期和时间
    now = datetime.now().strftime('%Y%m%d%H%M%S')

    # 新建 Excel 文件
    new_workbook = openpyxl.Workbook()

    # 获取源工作表
    source_sheet = workbook[source_sheet_name]

    # 创建新工作表并复制数据
    new_workbook.create_sheet(source_sheet_name)
    new_sheet = new_workbook[source_sheet_name]

    # 复制数据
    for row in source_sheet.iter_rows(values_only=True):
        new_sheet.append(row)

    # 删除默认创建的 Sheet
    del new_workbook['Sheet']

    # 保存新文件
    new_file_path = f'2026届13班座位表-{now}.xlsx'
    new_workbook.save(new_file_path)

    # 返回新文件路径
    return new_file_path


def remove_empty_rows(sheet):
    # 获取当前行数
    max_row = sheet.max_row

    # 从最后一行开始向前遍历，删除空行
    for row_num in range(max_row, 0, -1):
        if all(cell.value is None for cell in sheet[row_num]):
            sheet.delete_rows(row_num)


def main():
    # 读取“男生”工作表的第二列数据
    sheet_name_male = "男生"
    column_index_male = 2

    try:
        # 加载 Excel 文件
        workbook = openpyxl.load_workbook('2026届13班座位表源.xlsx')

        # 获取随机抽取的数据及其所在行号
        used_rows = []  # 用于记录已使用的行索引
        while True:
            random_row = random.randint(1, workbook[sheet_name_male].max_row)
            if random_row not in used_rows:
                break

        male_cell_value = workbook[sheet_name_male].cell(row=random_row, column=column_index_male).value
        used_rows.append(random_row)  # 记录已使用的行索引

        # 将随机抽取的数据填充到“输出座位表”的第4行第7列
        target_sheet_name_output = "输出座位表"
        target_row = 4
        target_column = 7

        assign_data_to_target_sheet(workbook, '2026届13班座位表源.xlsx', male_cell_value, target_sheet_name_output,
                                    target_row, target_column)
        print(f"Data has been assigned to '{target_sheet_name_output}' at row {target_row}, column {target_column}.")

        # 更新“男生_output”工作表，并记录已使用的数据
        output_sheet_name_male = "男生_output"

        # 获取原始工作表
        original_sheet = workbook[sheet_name_male]

        # 创建输出工作表
        if output_sheet_name_male not in workbook.sheetnames:
            workbook.create_sheet(output_sheet_name_male)
        output_sheet = workbook[output_sheet_name_male]

        # 清空输出工作表
        for row in output_sheet.iter_rows(min_row=1, max_col=1, max_row=original_sheet.max_row):
            for cell in row:
                cell.value = None

        # 将剩余的数据写入输出工作表
        for index, row_num in enumerate(range(1, original_sheet.max_row + 1)):
            if row_num not in used_rows:
                value = original_sheet.cell(row=row_num, column=column_index_male).value
                output_sheet.cell(row=index + 1, column=1).value = value

        # 保存更改后的 Excel 文件
        workbook.save('2026届13班座位表源.xlsx')
        print(f"Remaining data from '男生' sheet has been written to '{output_sheet_name_male}' in column 1.")

        # 获取“男生_output”工作表的数据
        output_sheet = workbook[output_sheet_name_male]
        output_data = [output_sheet.cell(row=row, column=1).value for row in range(1, output_sheet.max_row + 1)]

        # 删除“男生_output”工作表中的空行
        remove_empty_rows(output_sheet)
        workbook.save('2026届13班座位表源.xlsx')
        print("Empty rows have been removed from '男生_output'.")

        # 将“男生_output”工作表的第一列数据随机填充到“男生同桌”工作表的第二列第1~20行和第三列第1~20行
        target_sheet_name_male_partner = "男生同桌"
        target_columns_male = [2, 3]  # 目标列：第2列和第3列
        target_rows_male = list(range(1, 21))  # 目标行：第1行到第20行

        # 随机打乱数据
        random.shuffle(output_data)

        # 分配数据
        assign_output_data_to_target_sheet(workbook, '2026届13班座位表源.xlsx', output_sheet_name_male,
                                           target_sheet_name_male_partner, target_columns_male, target_rows_male)
        print(f"Data from '{output_sheet_name_male}' has been randomly assigned to '{target_sheet_name_male_partner}'.")

        # 读取“女生”工作表的第二列数据，并将其随机分配到“女生同桌”工作表的第2列和第3列的第1行到第6行
        sheet_name_female = "女生"
        column_index_female = 2

        # 获取“女生”工作表的数据
        female_sheet = workbook[sheet_name_female]
        female_data = [female_sheet.cell(row=row, column=column_index_female).value for row in
                       range(1, female_sheet.max_row + 1)]

        # 随机打乱数据
        random.shuffle(female_data)

        # 分配数据到“女生同桌”工作表
        target_sheet_name_female_partner = "女生同桌"
        target_columns_female = [2, 3]
        target_rows_female = list(range(1, 7))

        # 分配数据
        for i, value in enumerate(female_data[:len(target_columns_female) * len(target_rows_female)]):
            row_index = i // len(target_columns_female) + 1
            col_index = i % len(target_columns_female) + 1
            workbook[target_sheet_name_female_partner].cell(row=row_index,
                                                            column=target_columns_female[col_index - 1]).value = value

        # 保存更改后的 Excel 文件
        workbook.save('2026届13班座位表源.xlsx')
        print(f"Data from '女生' sheet has been randomly assigned to '{target_sheet_name_female_partner}'.")

        # 将“男生同桌”和“女生同桌”工作表中的数据随机分配到“座位表（按同桌）”工作表的第1行到第26行中
        partner_sheets = ["男生同桌", "女生同桌"]
        output_sheet_name_seats = "座位表（按同桌）"
        total_rows = 26

        assign_partner_data_to_output_sheet(workbook, '2026届13班座位表源.xlsx', partner_sheets,
                                            output_sheet_name_seats, total_rows)
        print(f"Data from '男生同桌' and '女生同桌' sheets has been randomly assigned to '{output_sheet_name_seats}'.")

        # 将“座位表（按同桌）”工作表中的数据复制到“输出座位表”工作表中
        copy_seats_to_output_sheet(workbook, '2026届13班座位表源.xlsx', output_sheet_name_seats,
                                   target_sheet_name_output)
        print(f"Data from '{output_sheet_name_seats}' has been copied to '{target_sheet_name_output}'.")

    except KeyError as e:
        print(e)
    except Exception as e:
        print(f"An error occurred: {e}")

    # 在主函数内部调用创建新文件的函数
    new_file_path = create_new_file_with_output_sheet(workbook, target_sheet_name_output)

    # 等待用户按回车键后打开新文件
    print("Press Enter to open the new file...")
    input()

    # 根据操作系统选择打开文件的方式
    if os.name == 'nt':  # Windows
        os.startfile(new_file_path)
    else:
        opener = "open" if platform.system() == "Darwin" else "xdg-open"
        subprocess.call([opener, new_file_path])


if __name__ == "__main__":
    main()